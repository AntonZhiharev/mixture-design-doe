"""
Utility functions for mixture designs
"""

import numpy as np
import pandas as pd
from typing import List, Tuple, Dict, Optional, Union
import itertools
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def validate_proportion_bounds(component_bounds: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    """
    Validate proportion bounds (0-1 range)
    
    Parameters:
    -----------
    component_bounds : List[Tuple[float, float]]
        List of (min, max) tuples for each component
        
    Returns:
    --------
    List[Tuple[float, float]] : Validated bounds
    
    Raises:
    -------
    ValueError : If bounds are invalid
    """
    validated_bounds = []
    
    for i, bounds_tuple in enumerate(component_bounds):
        try:
            lower, upper = bounds_tuple
            
            # Automatically fix swapped bounds (if lower > upper)
            if lower > upper:
                print(f"Warning: Swapped bounds detected for component {i+1}. "
                      f"Automatically reordering ({lower}, {upper}) to ({upper}, {lower})")
                lower, upper = upper, lower
            
            # Now validate the corrected bounds
            if lower < 0:
                raise ValueError(f"Invalid proportion bounds for component {i+1}: Lower bound {lower} is negative")
            if upper > 1:
                raise ValueError(f"Invalid proportion bounds for component {i+1}: Upper bound {upper} exceeds 1")
                
            validated_bounds.append((lower, upper))
        except ValueError as val_error:
            raise val_error
        except Exception as val_unexpected:
            raise ValueError(f"Cannot process bounds[{i}]: {bounds_tuple}") from val_unexpected
    
    # Check if bounds are feasible (sum of lower bounds <= 1, sum of upper bounds >= 1)
    try:
        sum_lower = sum(bound[0] for bound in validated_bounds)
        sum_upper = sum(bound[1] for bound in validated_bounds)
        
        if sum_lower > 1:
            print(f"Warning: Sum of lower bounds ({sum_lower:.4f}) exceeds 1. "
                  f"This may result in an infeasible mixture space or limit the design space.")
        if sum_upper < 1:
            print(f"Warning: Sum of upper bounds ({sum_upper:.4f}) is less than 1. "
                  f"This may result in an infeasible mixture space or require component scaling.")
            
    except Exception as sum_error:
        raise ValueError(f"Error validating bounds feasibility: {validated_bounds}") from sum_error
    
    return validated_bounds

def validate_parts_bounds(component_bounds: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    """
    Validate parts bounds (non-negative values)
    
    Parameters:
    -----------
    component_bounds : List[Tuple[float, float]]
        List of (min, max) tuples for each component in parts
        
    Returns:
    --------
    List[Tuple[float, float]] : Validated bounds
    
    Raises:
    -------
    ValueError : If bounds are invalid
    """
    validated_bounds = []
    
    for i, bounds_tuple in enumerate(component_bounds):
        try:
            lower, upper = bounds_tuple
            
            # Automatically fix swapped bounds (if lower > upper)
            if lower > upper:
                print(f"Warning: Swapped bounds detected for component {i+1}. "
                      f"Automatically reordering ({lower}, {upper}) to ({upper}, {lower})")
                lower, upper = upper, lower
            
            # Now validate the corrected bounds
            if lower < 0:
                raise ValueError(f"Invalid parts bounds for component {i+1}: ({lower}, {upper}). "
                                 f"Lower bound must be non-negative.")
                
            validated_bounds.append((lower, upper))
        except Exception as parts_error:
            raise ValueError(f"Cannot process parts bounds[{i}]: {bounds_tuple}") from parts_error
    
    return validated_bounds

def convert_parts_to_proportions(component_bounds_parts: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    """
    Convert bounds from parts to proportions
    
    Parameters:
    -----------
    component_bounds_parts : List[Tuple[float, float]]
        List of (min, max) tuples for each component in parts
        
    Returns:
    --------
    List[Tuple[float, float]] : Bounds converted to proportions
    """
    # Calculate total parts
    total_parts = sum(bound[1] for bound in component_bounds_parts)
    
    # Convert to proportions
    component_bounds_props = []
    for min_parts, max_parts in component_bounds_parts:
        min_prop = min_parts / total_parts
        max_prop = max_parts / total_parts
        component_bounds_props.append((min_prop, max_prop))
    
    return component_bounds_props

def check_bounds(point: List[float], component_bounds: List[Tuple[float, float]]) -> bool:
    """
    Check if a point satisfies component bounds
    
    Parameters:
    -----------
    point : List[float]
        Point to check
    component_bounds : List[Tuple[float, float]]
        List of (min, max) tuples for each component
        
    Returns:
    --------
    bool : True if point satisfies bounds, False otherwise
    """
    if len(point) != len(component_bounds):
        return False
    
    # Check if point sums to 1 (within tolerance)
    # Use a more lenient tolerance for sum check
    if not np.isclose(sum(point), 1.0, atol=1e-4):
        # Try to normalize the point if it's close enough
        point_sum = sum(point)
        if 0.9 < point_sum < 1.1:  # If within 10% of 1.0
            # Point can be normalized
            return True
        return False
    
    # Check if point satisfies bounds with a more lenient tolerance
    for i, (lower, upper) in enumerate(component_bounds):
        # Use a more lenient tolerance for bounds checking
        if point[i] < lower - 1e-4 or point[i] > upper + 1e-4:
            # Special case: if the bound is very close to 0 or 1
            if (lower <= 1e-4 and point[i] >= 0) or (upper >= 1.0 - 1e-4 and point[i] <= 1.0):
                continue
            return False
    
    return True

def generate_emphasis_point(bounds: List[Tuple[float, float]]) -> np.ndarray:
    """
    Generate a point that emphasizes one component at a time
    Used by extreme vertices design to create diverse points
    
    Parameters:
    -----------
    bounds : List[Tuple[float, float]]
        List of (min, max) tuples for each component
        
    Returns:
    --------
    np.ndarray : Point with emphasis on one component
    """
    n_components = len(bounds)
    
    # Randomly select one component to emphasize
    emphasis_idx = np.random.randint(0, n_components)
    
    # Create point with emphasis on selected component
    point = np.zeros(n_components)
    
    # Set all components to their lower bounds
    for i in range(n_components):
        if i == emphasis_idx:
            # Emphasized component gets a value closer to its upper bound
            lb, ub = bounds[i]
            # Use a value between 60-90% of the range
            emphasis_factor = np.random.uniform(0.6, 0.9)
            point[i] = lb + emphasis_factor * (ub - lb)
        else:
            # Other components get values closer to their lower bounds
            lb, ub = bounds[i]
            # Use a value between 10-40% of the range
            background_factor = np.random.uniform(0.1, 0.4)
            point[i] = lb + background_factor * (ub - lb)
    
    return point

def generate_boundary_biased_point(component_bounds: List[Tuple[float, float]]) -> np.ndarray:
    """
    Generate a point that is biased toward the boundaries of the feasible region
    Used to improve diversity in candidate point generation
    
    Parameters:
    -----------
    component_bounds : List[Tuple[float, float]]
        List of (min, max) tuples for each component
        
    Returns:
    --------
    np.ndarray : Point biased toward boundaries
    """
    n_components = len(component_bounds)
    point = np.zeros(n_components)
    
    # For each component, decide whether to use a value near the boundary
    for i in range(n_components):
        lb, ub = component_bounds[i]
        
        # 60% chance of using a boundary value
        if np.random.random() < 0.6:
            # 50% chance of using lower bound, 50% chance of using upper bound
            if np.random.random() < 0.5:
                # Use value near lower bound (within 20% of range)
                range_val = ub - lb
                point[i] = lb + np.random.uniform(0, 0.2) * range_val
            else:
                # Use value near upper bound (within 20% of range)
                range_val = ub - lb
                point[i] = ub - np.random.uniform(0, 0.2) * range_val
        else:
            # Use random value in middle range
            range_val = ub - lb
            point[i] = lb + np.random.uniform(0.2, 0.8) * range_val
    
    return point

def generate_binary_mixture(n_components: int) -> np.ndarray:
    """
    Generate a binary mixture (only two components have non-zero values)
    Used to improve diversity in candidate point generation
    
    Parameters:
    -----------
    n_components : int
        Number of components
        
    Returns:
    --------
    np.ndarray : Binary mixture point
    """
    point = np.zeros(n_components)
    
    # Select two components randomly
    if n_components >= 2:
        i, j = np.random.choice(range(n_components), 2, replace=False)
        
        # Random ratio between the two components
        ratio = np.random.random()
        
        # Set values for the two components
        point[i] = ratio
        point[j] = 1.0 - ratio
    else:
        # Fallback for single component case
        point[0] = 1.0
    
    return point

def adjust_largest_more(point: np.ndarray, indices: List[int], delta: float) -> List[float]:
    """
    Adjustment strategy that reduces the largest component more when increasing another component
    This helps explore more diverse designs by creating more extreme points
    
    Parameters:
    -----------
    point : np.ndarray
        Current point
    indices : List[int]
        Indices of components to adjust
    delta : float
        Amount to adjust by (positive means increasing a component, negative means decreasing)
        
    Returns:
    --------
    List[float] : Adjusted values for the components in indices
    """
    if not indices:
        return []
        
    # Find the largest component
    values = [point[i] for i in indices]
    largest_idx = indices[np.argmax(values)]
    largest_val = point[largest_idx]
    
    # Calculate sum of all components to adjust
    total = sum(values)
    
    # If delta is positive, we're increasing another component, so reduce others
    if delta > 0:
        # If largest component is big enough, take most from it
        if largest_val > 0.5 * total:
            # Take 70% from largest, 30% proportionally from others
            largest_reduction = delta * 0.7
            other_reduction = delta * 0.3
            
            # Adjust values
            adjusted = []
            for i in indices:
                if i == largest_idx:
                    # Reduce largest component more
                    new_val = point[i] - largest_reduction
                    adjusted.append(max(0, new_val))
                else:
                    # Reduce other components proportionally
                    other_sum = total - largest_val
                    if other_sum > 0:
                        reduction_factor = point[i] / other_sum
                        new_val = point[i] - (other_reduction * reduction_factor)
                        adjusted.append(max(0, new_val))
                    else:
                        adjusted.append(point[i])
        else:
            # Reduce all proportionally
            adjusted = []
            for i in indices:
                reduction_factor = point[i] / total
                new_val = point[i] - (delta * reduction_factor)
                adjusted.append(max(0, new_val))
    else:
        # If delta is negative, we're decreasing another component, so increase others
        # Increase all proportionally
        adjusted = []
        for i in indices:
            increase_factor = point[i] / total
            new_val = point[i] - (delta * increase_factor)  # delta is negative, so this increases
            adjusted.append(new_val)
    
    return adjusted

def select_diverse_subset(design: np.ndarray, n_select: int) -> np.ndarray:
    """
    Select diverse subset using maximin distance criterion
    
    Parameters:
    -----------
    design : np.ndarray
        Design matrix
    n_select : int
        Number of points to select
        
    Returns:
    --------
    np.ndarray : Selected subset of design points
    """
    from sklearn.metrics import pairwise_distances
    
    # Calculate pairwise distances
    distances = pairwise_distances(design)
    
    # Start with two most distant points
    i, j = np.unravel_index(np.argmax(distances), distances.shape)
    selected = [i, j]
    
    # Iteratively add points that maximize minimum distance
    while len(selected) < n_select:
        remaining = [idx for idx in range(len(design)) if idx not in selected]
        if not remaining:
            break
        
        best_idx = remaining[0]
        best_min_dist = 0
        
        for idx in remaining:
            # Find minimum distance to selected points
            min_dist = np.min([distances[idx, s] for s in selected])
            if min_dist > best_min_dist:
                best_min_dist = min_dist
                best_idx = idx
        
        selected.append(best_idx)
    
    return design[selected]


def create_enhanced_excel_export(design_df: pd.DataFrame, 
                                component_names: List[str],
                                use_parts_mode: bool = False,
                                parts_design: Optional[np.ndarray] = None,
                                batch_sizes: Optional[List[float]] = None,
                                filename: Optional[str] = None) -> Union[bytes, None]:
    """
    Create enhanced Excel export with run numbers, percentages, and kg quantities.
    
    Parameters:
    -----------
    design_df : pd.DataFrame
        Design matrix DataFrame (proportions)
    component_names : List[str]
        Names of components
    use_parts_mode : bool, default False
        Whether parts mode is being used
    parts_design : np.ndarray, optional
        Parts design matrix (if parts mode is used)
    batch_sizes : List[float], optional
        List of batch sizes for manufacturing worksheets
    filename : str, optional
        If provided, saves to file. If None, returns bytes.
        
    Returns:
    --------
    bytes or None
        Excel file as bytes if filename is None, otherwise saves to file
    """
    # Create Excel writer
    if filename:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
    else:
        excel_buffer = io.BytesIO()
        writer = pd.ExcelWriter(excel_buffer, engine='openpyxl')
    
    try:
        # Sheet 1: Design Matrix with Run Numbers, Percentages, and Parts (if applicable)
        if use_parts_mode and parts_design is not None:
            enhanced_design_df = add_run_numbers_with_parts(design_df, component_names, parts_design)
        else:
            enhanced_design_df = add_run_numbers(design_df, component_names)
        
        enhanced_design_df.to_excel(writer, sheet_name='Design_Matrix', index=False)
        format_design_sheet(writer.sheets['Design_Matrix'], enhanced_design_df)
        
        # Sheet 2: Manufacturing Worksheets (if batch sizes provided)
        if use_parts_mode and parts_design is not None and batch_sizes:
            for i, batch_size in enumerate(batch_sizes):
                sheet_name = f'Manufacturing_{batch_size}kg'
                manufacturing_df = create_manufacturing_worksheet(
                    parts_design, component_names, batch_size
                )
                manufacturing_df.to_excel(writer, sheet_name=sheet_name, index=False)
                format_manufacturing_sheet(writer.sheets[sheet_name], manufacturing_df)
                
                # Add material summary
                summary_sheet_name = f'Materials_{batch_size}kg'
                summary_df = create_material_summary(
                    parts_design, component_names, batch_size
                )
                summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
                format_summary_sheet(writer.sheets[summary_sheet_name], summary_df)
        
        # Save and return
        writer.close()
        
        if filename:
            print(f"✅ Enhanced Excel file saved to {filename}")
            return None
        else:
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
    except Exception as e:
        writer.close()
        raise Exception(f"Error creating Excel export: {str(e)}")


def add_run_numbers(design_df: pd.DataFrame, component_names: List[str]) -> pd.DataFrame:
    """
    Add run numbers to design DataFrame with percentages only (no duplication).
    
    Parameters:
    -----------
    design_df : pd.DataFrame
        Original design DataFrame
    component_names : List[str]
        Names of components
        
    Returns:
    --------
    pd.DataFrame
        Enhanced DataFrame with run numbers and percentages
    """
    enhanced_df = pd.DataFrame()
    
    # Add run numbers as first column
    enhanced_df['Run_Number'] = range(1, len(design_df) + 1)
    
    # Handle column mapping - design_df might have different column names than component_names
    available_columns = design_df.columns.tolist()
    n_available = len(available_columns)
    
    # Only process components that have corresponding data columns
    # Ensure we don't create empty columns for non-existent components
    effective_component_names = component_names[:n_available] if len(component_names) > n_available else component_names
    
    # Add percentage columns only for components that have data
    for i, comp_name in enumerate(effective_component_names):
        if comp_name in design_df.columns:
            # Direct match
            enhanced_df[f'{comp_name}_%'] = (design_df[comp_name] * 100).round(1)
        elif i < n_available:
            # Use positional matching if name doesn't match
            enhanced_df[f'{comp_name}_%'] = (design_df.iloc[:, i] * 100).round(1)
        # Removed the else clause that created empty columns with zeros
    
    # Add verification columns
    percentage_cols = [f'{name}_%' for name in effective_component_names if f'{name}_%' in enhanced_df.columns]
    enhanced_df['Total_%'] = enhanced_df[percentage_cols].sum(axis=1)
    enhanced_df['Sum_Check'] = enhanced_df['Total_%'].apply(lambda x: '✓' if abs(x - 100) < 0.1 else '✗')
    
    return enhanced_df


def add_run_numbers_with_parts(design_df: pd.DataFrame, component_names: List[str], parts_design: np.ndarray) -> pd.DataFrame:
    """
    Add run numbers to design DataFrame with both percentages and parts quantities.
    
    Parameters:
    -----------
    design_df : pd.DataFrame
        Original design DataFrame (proportions)
    component_names : List[str]
        Names of components
    parts_design : np.ndarray
        Parts design matrix
        
    Returns:
    --------
    pd.DataFrame
        Enhanced DataFrame with run numbers, percentages, and parts quantities
    """
    enhanced_df = pd.DataFrame()
    
    # Add run numbers as first column
    enhanced_df['Run_Number'] = range(1, len(design_df) + 1)
    
    # Ensure we don't exceed the number of available columns in parts_design
    n_parts_columns = parts_design.shape[1]
    effective_component_names = component_names[:n_parts_columns] if len(component_names) > n_parts_columns else component_names
    
    # Add percentage columns
    for i, comp_name in enumerate(effective_component_names):
        if comp_name in design_df.columns:
            # Direct match
            enhanced_df[f'{comp_name}_%'] = (design_df[comp_name] * 100).round(1)
        elif i < len(design_df.columns):
            # Use positional matching if name doesn't match
            enhanced_df[f'{comp_name}_%'] = (design_df.iloc[:, i] * 100).round(1)
    
    # Add parts quantities (only for components that have corresponding data)
    for i, comp_name in enumerate(effective_component_names):
        enhanced_df[f'{comp_name}_parts'] = parts_design[:, i].round(4)
    
    # Add verification columns
    enhanced_df['Total_%'] = enhanced_df[[f'{name}_%' for name in effective_component_names if f'{name}_%' in enhanced_df.columns]].sum(axis=1)
    enhanced_df['Total_parts'] = parts_design.sum(axis=1).round(4)
    enhanced_df['Sum_Check'] = enhanced_df['Total_%'].apply(lambda x: '✓' if abs(x - 100) < 0.1 else '✗')
    
    return enhanced_df


def create_parts_dataframe(parts_design: np.ndarray, component_names: List[str]) -> pd.DataFrame:
    """
    Create parts DataFrame with run numbers.
    
    Parameters:
    -----------
    parts_design : np.ndarray
        Parts design matrix
    component_names : List[str]
        Names of components
        
    Returns:
    --------
    pd.DataFrame
        Parts DataFrame with run numbers
    """
    parts_df = pd.DataFrame(parts_design, columns=component_names)
    
    # Add run numbers
    parts_df.insert(0, 'Run_Number', range(1, len(parts_design) + 1))
    
    # Add total parts column
    parts_df['Total_Parts'] = parts_design.sum(axis=1)
    
    return parts_df


def create_manufacturing_worksheet(parts_design: np.ndarray, 
                                 component_names: List[str], 
                                 batch_size: float) -> pd.DataFrame:
    """
    Create manufacturing worksheet for specific batch size.
    
    Parameters:
    -----------
    parts_design : np.ndarray
        Parts design matrix
    component_names : List[str]
        Names of components
    batch_size : float
        Target batch size in kg
        
    Returns:
    --------
    pd.DataFrame
        Manufacturing worksheet
    """
    # Calculate actual quantities for this batch size
    total_parts_per_run = parts_design.sum(axis=1, keepdims=True)
    actual_quantities = parts_design * batch_size / total_parts_per_run
    
    # Ensure we don't exceed the number of available columns in parts_design
    n_parts_columns = parts_design.shape[1]
    effective_component_names = component_names[:n_parts_columns] if len(component_names) > n_parts_columns else component_names
    
    # Create worksheet DataFrame
    worksheet_df = pd.DataFrame()
    
    # Add run identification
    worksheet_df['Run_Number'] = range(1, len(parts_design) + 1)
    worksheet_df['Experiment_ID'] = [f'EXP_{i+1:03d}' for i in range(len(parts_design))]
    
    # Add all percentage columns first (only for components that have data)
    for j, comp_name in enumerate(effective_component_names):
        worksheet_df[f'{comp_name}_%'] = (parts_design[:, j] / total_parts_per_run.flatten() * 100).round(2)
    
    # Then add all kg quantity columns (only for components that have data)
    for j, comp_name in enumerate(effective_component_names):
        worksheet_df[f'{comp_name}_kg'] = actual_quantities[:, j].round(4)
    
    # Add totals and verification
    worksheet_df['Total_kg'] = actual_quantities.sum(axis=1).round(4)
    worksheet_df['Weight_Check'] = ['✓' if abs(total - batch_size) < 1e-3 else '✗' 
                                   for total in worksheet_df['Total_kg']]
    
    return worksheet_df


def create_material_summary(parts_design: np.ndarray, 
                           component_names: List[str], 
                           batch_size: float) -> pd.DataFrame:
    """
    Create material requirements summary for specific batch size.
    
    Parameters:
    -----------
    parts_design : np.ndarray
        Parts design matrix
    component_names : List[str]
        Names of components
    batch_size : float
        Target batch size in kg
        
    Returns:
    --------
    pd.DataFrame
        Material requirements summary
    """
    # Calculate actual quantities for this batch size
    total_parts_per_run = parts_design.sum(axis=1, keepdims=True)
    actual_quantities = parts_design * batch_size / total_parts_per_run
    
    # Ensure we don't exceed the number of available columns in parts_design
    n_parts_columns = parts_design.shape[1]
    effective_component_names = component_names[:n_parts_columns] if len(component_names) > n_parts_columns else component_names
    
    # Create summary using only components that have data
    total_materials = actual_quantities.sum(axis=0)
    
    summary_df = pd.DataFrame({
        'Component': effective_component_names,
        'Total_Required_kg': total_materials.round(3),
        'Min_per_Run_kg': actual_quantities.min(axis=0).round(3),
        'Max_per_Run_kg': actual_quantities.max(axis=0).round(3),
        'Order_with_20%_Buffer_kg': (total_materials * 1.2).round(3)
    })
    
    # Add summary rows
    totals_row = pd.DataFrame({
        'Component': ['TOTAL'],
        'Total_Required_kg': [total_materials.sum().round(3)],
        'Min_per_Run_kg': [actual_quantities.sum(axis=1).min().round(3)],
        'Max_per_Run_kg': [actual_quantities.sum(axis=1).max().round(3)],
        'Order_with_20%_Buffer_kg': [(total_materials.sum() * 1.2).round(3)]
    })
    
    summary_df = pd.concat([summary_df, totals_row], ignore_index=True)
    
    return summary_df


def format_design_sheet(worksheet, df: pd.DataFrame):
    """Format the design matrix sheet."""
    try:
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Warning: Could not format design sheet: {e}")


def format_parts_sheet(worksheet, df: pd.DataFrame):
    """Format the parts design sheet."""
    try:
        # Header formatting
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Warning: Could not format parts sheet: {e}")


def format_manufacturing_sheet(worksheet, df: pd.DataFrame):
    """Format the manufacturing worksheet."""
    try:
        # Header formatting
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        header_font = Font(bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Highlight weight check column
        if 'Weight_Check' in df.columns:
            check_col = df.columns.get_loc('Weight_Check') + 1
            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=check_col)
                if cell.value == '✓':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Warning: Could not format manufacturing sheet: {e}")


def format_summary_sheet(worksheet, df: pd.DataFrame):
    """Format the material summary sheet."""
    try:
        # Header formatting
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Highlight total row
        if len(df) > 0:
            total_row = len(df) + 1
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=total_row, column=col_num)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        print(f"Warning: Could not format summary sheet: {e}")


def clean_numerical_precision(data: Union[np.ndarray, pd.DataFrame, List], 
                             zero_threshold: float = None, 
                             one_threshold: float = None,
                             preserve_mixture_constraint: bool = True,
                             data_scale: str = "auto") -> Union[np.ndarray, pd.DataFrame, List]:
    """
    Clean up numerical precision issues in mixture design data.
    
    Converts near-zero values to exact 0.0 and near-maximum values to exact maximum.
    This prevents display issues like 0.9998 or 0.0002 in mixture designs.
    For mixture data, preserves the sum constraint by normalizing after cleanup.
    
    Parameters:
    -----------
    data : np.ndarray, pd.DataFrame, or List
        Data to clean up
    zero_threshold : float, optional
        Values below this threshold become 0.0. Auto-detected if None.
    one_threshold : float, optional
        Values above (maximum - one_threshold) become maximum. Auto-detected if None.
    preserve_mixture_constraint : bool, default True
        If True, normalize each row to preserve sum constraint (for mixture data)
    data_scale : str, default "auto"
        Data scale: "auto", "proportions" (0-1), "parts" (0-100), or "custom"
        
    Returns:
    --------
    Same type as input
        Data with cleaned numerical precision
    """
    
    def detect_data_scale(data_values):
        """Auto-detect if data is proportions (0-1) or parts (0-100+)"""
        if isinstance(data_values, (list, tuple)):
            data_values = np.array(data_values)
        
        if hasattr(data_values, 'values'):  # DataFrame
            data_values = data_values.values
        
        max_val = np.max(data_values)
        min_val = np.min(data_values)
        
        if max_val <= 1.1 and min_val >= -0.1:
            return "proportions"  # 0-1 range
        elif max_val <= 110 and min_val >= -1:
            return "parts"  # 0-100 range (with some tolerance)
        else:
            return "custom"  # Other scale
    
    def get_thresholds(scale, zero_threshold, one_threshold):
        """Get appropriate thresholds for the data scale"""
        if zero_threshold is not None and one_threshold is not None:
            return zero_threshold, one_threshold
        
        if scale == "proportions":
            return 1e-3, 1e-3  # 0.001 for 0-1 scale
        elif scale == "parts":
            return 0.1, 0.1    # 0.1 for 0-100 scale
        else:
            return 1e-3, 1e-3  # Default fallback
    
    def get_maximum_value(scale):
        """Get the maximum value for the scale"""
        if scale == "proportions":
            return 1.0
        elif scale == "parts":
            return 100.0
        else:
            return 1.0  # Default fallback
    # Auto-detect scale if needed
    if data_scale == "auto":
        data_scale = detect_data_scale(data)
    
    # Get appropriate thresholds for the scale
    zero_thresh, one_thresh = get_thresholds(data_scale, zero_threshold, one_threshold)
    max_value = get_maximum_value(data_scale)
    
    def normalize_mixture_row(row_data, scale):
        """Normalize a row to preserve sum constraint while keeping cleaned zeros"""
        if isinstance(row_data, (list, tuple)):
            row_array = np.array(row_data)
        else:
            row_array = row_data.copy()
        
        # Get target sum based on scale
        target_sum = max_value if scale == "parts" else 1.0
        
        # Only normalize if the row doesn't already sum to target
        row_sum = np.sum(row_array)
        if not np.isclose(row_sum, target_sum, atol=1e-10):
            # Find non-zero elements for normalization
            non_zero_mask = row_array > 0
            if np.any(non_zero_mask):
                # Normalize only non-zero elements to preserve exact zeros
                row_array[non_zero_mask] = row_array[non_zero_mask] * target_sum / row_sum
        
        return row_array
    
    if isinstance(data, pd.DataFrame):
        # Process DataFrame
        cleaned_df = data.copy()
        for col in cleaned_df.columns:
            if cleaned_df[col].dtype in ['float64', 'float32']:
                # Clean numerical precision for float columns
                values = cleaned_df[col].values
                values[values < zero_thresh] = 0.0
                values[values > max_value - one_thresh] = max_value
                cleaned_df[col] = values
        
        # Normalize each row to preserve mixture constraint
        if preserve_mixture_constraint:
            for idx in range(len(cleaned_df)):
                row_values = cleaned_df.iloc[idx].values
                normalized_row = normalize_mixture_row(row_values, data_scale)
                cleaned_df.iloc[idx] = normalized_row
        
        return cleaned_df
    
    elif isinstance(data, np.ndarray):
        # Process numpy array
        cleaned_data = data.copy()
        cleaned_data[cleaned_data < zero_thresh] = 0.0
        cleaned_data[cleaned_data > max_value - one_thresh] = max_value
        
        # Normalize each row to preserve mixture constraint
        if preserve_mixture_constraint and cleaned_data.ndim == 2:
            for i in range(cleaned_data.shape[0]):
                cleaned_data[i] = normalize_mixture_row(cleaned_data[i], data_scale)
        
        return cleaned_data
    
    elif isinstance(data, list):
        # Process list (potentially nested)
        if data and isinstance(data[0], (list, tuple)):
            # Nested list/tuple
            cleaned_data = []
            for row in data:
                cleaned_row = []
                for val in row:
                    if isinstance(val, (int, float)):
                        if val < zero_thresh:
                            cleaned_row.append(0.0)
                        elif val > max_value - one_thresh:
                            cleaned_row.append(max_value)
                        else:
                            cleaned_row.append(val)
                    else:
                        cleaned_row.append(val)
                
                # Normalize row if preserving mixture constraint
                if preserve_mixture_constraint:
                    cleaned_row = normalize_mixture_row(cleaned_row, data_scale).tolist()
                
                cleaned_data.append(cleaned_row)
            return cleaned_data
        else:
            # Simple list
            cleaned_data = []
            for val in data:
                if isinstance(val, (int, float)):
                    if val < zero_thresh:
                        cleaned_data.append(0.0)
                    elif val > max_value - one_thresh:
                        cleaned_data.append(max_value)
                    else:
                        cleaned_data.append(val)
                else:
                    cleaned_data.append(val)
            return cleaned_data
    
    else:
        # Return as-is for other types
        return data


def ensure_sum_constraint(design_matrix: np.ndarray, tolerance: float = 1e-10) -> np.ndarray:
    """
    Ensure each row of the design matrix sums to 1.0 within tolerance.
    
    Parameters:
    -----------
    design_matrix : np.ndarray
        The design matrix to normalize
    tolerance : float
        Tolerance for sum constraint (default: 1e-10)
        
    Returns:
    --------
    np.ndarray
        Design matrix with rows normalized to sum to 1.0
    """
    # Calculate row sums
    row_sums = np.sum(design_matrix, axis=1)
    
    # Check which rows need normalization
    needs_normalization = np.abs(row_sums - 1.0) > tolerance
    
    if np.any(needs_normalization):
        # Normalize only rows that need it
        design_matrix[needs_normalization] = design_matrix[needs_normalization] / row_sums[needs_normalization, np.newaxis]
        
        # Warn user
        n_normalized = np.sum(needs_normalization)
        print(f"Warning: Normalized {n_normalized} rows to satisfy sum constraint. "
              f"Max deviation was {np.max(np.abs(row_sums - 1.0)):.2e}")
    
    return design_matrix
